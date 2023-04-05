import time
import selenium.webdriver as webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager 
from openpyxl.styles import Font
from datetime import date
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.utils.dataframe import dataframe_to_rows


urls = ['https://decrypt.co/news']

options = Options()
options.add_experimental_option("detach",True)
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.maximize_window()

SHEET_NAME = 'decrypt.co'

for url in urls:
    driver.get(url)
    time.sleep(5)

    tab_base_name = url.split('/')[-2]
    today = date.today()
    d1 = today.strftime("%d_%m_%Y")
    print(f'Scraping {url}...')
    xlwriter = pd.ExcelWriter('News.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay')
    existing_data = pd.read_excel('News.xlsx', engine='openpyxl', sheet_name=SHEET_NAME)
    workbook = xlwriter.book
    
    site = driver.page_source
    soup = BeautifulSoup(site,"html.parser")
    decrypt = []
    news_row = soup.find_all('div',class_="sc-50880c4b-0")
    
    headers = ["source","headline","content","genre","date"]
    
    for news in news_row:
        source_url = "https://"+tab_base_name+news.find('a',class_="block").get('href')
        headline = news.find('h2',class_="mb-2").text
        content = news.find('p',class_="mb-3").text
        genre = news.find('span',class_="hover:underline").text
        date_time = news.find('time').text  
        row_data = [source_url, headline, content, genre, date_time]
        if row_data not in existing_data.values.tolist():
            decrypt.append(row_data)
        else:
            pass

    df = pd.DataFrame(decrypt,columns=headers)
    sheet = workbook[SHEET_NAME]
    next_row = sheet.max_row + 1
    for row in dataframe_to_rows(df, index=False, header=False):
        for col, value in enumerate(row, start=1):
            sheet.cell(row=next_row, column=col, value=value)
        next_row += 1

    font_style = Font(color='0000FF', underline='single')
    ws = workbook.active
    for row in range(2, ws.max_row+1):
        cell = 'A{}'.format(row)
        link = ws[cell].value
        if link:
            ws.cell(row=row, column=1).hyperlink = link
            ws.cell(row=row, column=1).font = font_style

    print(df)

xlwriter.save()
xlwriter.close()
driver.close()
